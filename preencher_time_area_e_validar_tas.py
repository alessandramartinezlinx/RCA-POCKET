#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Preenche automaticamente as colunas Time (L) e Área (M) na planilha RCA Pocket
e executa validação de TAs no GitHub Robot Framework.

Uso:
    python preencher_time_area_e_validar_tas.py
"""

import os
import sys
import json
import yaml
import openpyxl
from datetime import datetime

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================
CACHE_FILE = "data/issues_cache.json"
EXCEL_FILE = "RCA_Pocket.xlsx"
CONFIG_FILE = "rca_config.yaml"
SHEET_NAME = "📊 Dados"

COL_KEY = 1    # A
COL_TIME = 12  # L
COL_AREA = 13  # M

# =============================================================================
# FUNÇÕES AUXILIARES
# =============================================================================

def carregar_config():
    """Carrega configuração do rca_config.yaml"""
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def carregar_cache():
    """Carrega cache de issues do Jira"""
    if not os.path.exists(CACHE_FILE):
        print(f"❌ Cache não encontrado: {CACHE_FILE}")
        print("   Execute primeiro: python jira_client.py")
        return None
    
    with open(CACHE_FILE, 'r', encoding='utf-8') as f:
        cache = json.load(f)
    return cache.get('issues', [])

def mapear_labels_para_time_area(labels, config):
    """
    Mapeia labels para time e área usando a configuração.
    
    Args:
        labels: Lista de strings (labels do Jira)
        config: Dicionário com configuração do rca_config.yaml
    
    Returns:
        Tupla (time, area) ou (None, None) se não encontrado
    """
    if not labels:
        return None, None
    
    times = config.get('times', {})
    
    # Para cada time definido
    for time_nome, time_config in times.items():
        areas = time_config.get('areas', [])
        
        # Para cada área do time
        for area_config in areas:
            area_nome = area_config['nome']
            labels_jira = area_config.get('labels_jira', [])
            
            # Verifica se alguma label da issue bate com as labels_jira da área
            for label in labels:
                if label in labels_jira:
                    return time_nome, area_nome
    
    return None, None

def inferir_time_area_por_key(key):
    """
    Infere time/área baseado no padrão da key (fallback).
    
    Exemplos:
        - MODAJOI-XXXXX → poderia ser de qualquer time
        - SUP-XXX → Suprimentos
        - FFC-XXX → FFC
    """
    # Por enquanto, sem inferência específica
    # Pode ser expandido conforme padrões identificados
    return None, None

def preencher_time_area_planilha():
    """
    Preenche colunas Time (L) e Área (M) na planilha baseado no cache e config.
    """
    print("\n" + "="*70)
    print("  🔧 PREENCHIMENTO AUTOMÁTICO DE TIME E ÁREA")
    print("="*70)
    
    # Carregar dados
    print("\n📦 Carregando configuração e cache...")
    config = carregar_config()
    issues = carregar_cache()
    
    if not issues:
        return False
    
    print(f"   ✅ {len(issues)} issues no cache")
    
    # Criar índice de issues por key
    issues_by_key = {i['key']: i for i in issues}
    
    # Abrir planilha
    print(f"\n📄 Abrindo planilha: {EXCEL_FILE}...")
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Planilha não encontrada: {EXCEL_FILE}")
        print("   Execute primeiro: python generate_excel.py")
        return False
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    # Processar cada linha
    atualizadas = 0
    sem_mapeamento = []
    
    print(f"\n🔍 Processando issues...")
    for row_idx in range(2, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=COL_KEY)
        key = key_cell.value
        
        if not key:
            continue
        
        # Buscar issue no cache
        issue = issues_by_key.get(key)
        if not issue:
            print(f"   ⚠️  {key}: não encontrada no cache")
            continue
        
        # Tentar mapear labels
        labels = issue.get('labels', [])
        time, area = mapear_labels_para_time_area(labels, config)
        
        # Fallback: tentar inferir por padrão da key
        if not time:
            time, area = inferir_time_area_por_key(key)
        
        # Atualizar planilha
        time_cell = ws.cell(row=row_idx, column=COL_TIME)
        area_cell = ws.cell(row=row_idx, column=COL_AREA)
        
        # Só atualiza se encontrou mapeamento
        if time:
            time_cell.value = time
            area_cell.value = area
            atualizadas += 1
            print(f"   ✅ {key}: {time} > {area}")
        else:
            # Marca como não mapeada
            sem_mapeamento.append({
                'key': key,
                'labels': labels
            })
            print(f"   ⚠️  {key}: sem mapeamento (labels: {labels or 'vazio'})")
    
    # Salvar planilha
    print(f"\n💾 Salvando planilha...")
    wb.save(EXCEL_FILE)
    print(f"   ✅ Planilha atualizada")
    
    # Resumo
    print("\n" + "="*70)
    print("📊 RESUMO DO PREENCHIMENTO")
    print("="*70)
    print(f"  Total de issues na planilha: {ws.max_row - 1}")
    print(f"  ✅ Atualizadas com sucesso: {atualizadas}")
    print(f"  ⚠️  Sem mapeamento: {len(sem_mapeamento)}")
    
    if sem_mapeamento:
        print("\n⚠️  Issues sem mapeamento (adicione labels no Jira ou configure rca_config.yaml):")
        for item in sem_mapeamento[:10]:  # Mostrar só as primeiras 10
            print(f"     • {item['key']}: labels = {item['labels'] or '(vazio)'}")
        if len(sem_mapeamento) > 10:
            print(f"     ... e mais {len(sem_mapeamento) - 10}")
    
    print("="*70)
    return True

def executar_validacao_tas():
    """
    Executa validação de TAs após preencher time/área.
    """
    print("\n" + "="*70)
    print("  🤖 VALIDAÇÃO DE TESTES AUTOMATIZADOS")
    print("="*70)
    
    # Verificar token GitHub
    github_token = os.getenv("GITHUB_TOKEN")
    if not github_token:
        print("\n⚠️  Token GitHub não configurado!")
        print("   Configure GITHUB_TOKEN para validar TAs:")
        print("   1. Crie token em: https://github.com/settings/tokens")
        print("   2. PowerShell: $env:GITHUB_TOKEN = 'ghp_seu_token'")
        print("\n   Pulando validação de TAs...")
        return False
    
    # Verificar se script existe
    if not os.path.exists("validar_tas_planilha.py"):
        print("\n⚠️  Script validar_tas_planilha.py não encontrado!")
        return False
    
    # Executar validação
    print("\n🚀 Executando validação de TAs...")
    result = os.system(f"{sys.executable} validar_tas_planilha.py")
    
    return result == 0

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n╔══════════════════════════════════════════════════════════╗")
    print("║  RCA Pocket - Preenchimento Automático + Validação TAs  ║")
    print("╚══════════════════════════════════════════════════════════╝")
    
    # Etapa 1: Preencher Time e Área
    sucesso_time_area = preencher_time_area_planilha()
    
    if not sucesso_time_area:
        print("\n❌ Falha ao preencher Time/Área")
        return 1
    
    # Etapa 2: Validar TAs (se token configurado)
    print("\n" + "─"*70)
    resposta = input("\n🤖 Deseja executar validação de TAs agora? [S/n]: ").strip().lower()
    
    if resposta in ('', 's', 'sim', 'y', 'yes'):
        executar_validacao_tas()
    else:
        print("\n   Validação de TAs pulada.")
        print("   Execute manualmente: python validar_tas_planilha.py")
    
    print("\n✅ Processo concluído!")
    print("   Abra RCA_Pocket.xlsx para visualizar as alterações.\n")
    return 0

if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n\n⚠️  Processo cancelado pelo usuário")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Erro inesperado: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
