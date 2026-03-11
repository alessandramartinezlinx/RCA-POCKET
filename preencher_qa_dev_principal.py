#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Preenche automaticamente as colunas QA Principal (W) e Dev Principal (X)
na planilha RCA Pocket com base nas colunas Time (L) e Área (M).

Busca o mapeamento no arquivo rca_config.yaml.

Uso:
    python preencher_qa_dev_principal.py
"""

import yaml
import openpyxl
from typing import Dict, Tuple, Optional

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================
EXCEL_FILE = "RCA_Pocket.xlsx"
CONFIG_FILE = "rca_config.yaml"
SHEET_NAME = "📊 Dados"

COL_KEY = 1           # A - Key
COL_TIME = 12         # L - Time
COL_AREA = 13         # M - Área
COL_QA_PRINCIPAL = 23 # W - QA Principal
COL_DEV_PRINCIPAL = 24 # X - Dev Principal

# =============================================================================
# FUNÇÕES
# =============================================================================

def carregar_config() -> dict:
    """Carrega configuração do rca_config.yaml"""
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def criar_mapeamento_times(config: dict) -> Dict[Tuple[str, str], Dict[str, str]]:
    """
    Cria dicionário de mapeamento: (Time, Área) -> {qa_principal, dev_principal}
    
    Returns:
        {
            ("Suprimentos", "Entrada XML"): {"qa_principal": "Jenifer Lopes", "dev_principal": "Vinícius Souza Martins"},
            ("FFC", "NF-e"): {"qa_principal": "Jean Carlos", "dev_principal": "Flavia Doge"},
            ...
        }
    """
    mapeamento = {}
    
    times = config.get('times', {})
    for time_name, time_data in times.items():
        areas = time_data.get('areas', [])
        for area in areas:
            area_nome = area.get('nome', '')
            qa_principal = area.get('qa_principal', '')
            dev_principal = area.get('dev_principal', '')
            
            chave = (time_name, area_nome)
            mapeamento[chave] = {
                'qa_principal': qa_principal,
                'dev_principal': dev_principal
            }
    
    return mapeamento


def preencher_qa_dev():
    """Função principal de preenchimento"""
    print("🔧 PREENCHIMENTO DE QA E DEV PRINCIPAL")
    print("="*70)
    
    # 1. Carregar config
    print("\n📦 Carregando configuração...")
    config = carregar_config()
    mapeamento = criar_mapeamento_times(config)
    print(f"   ✅ {len(mapeamento)} combinações Time/Área mapeadas")
    
    # 2. Abrir planilha
    print(f"\n📄 Abrindo planilha {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    # 3. Processar issues
    print(f"\n🔍 Processando issues...")
    atualizadas = 0
    sem_mapeamento = []
    ja_preenchidas = 0
    
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, COL_KEY).value
        if not key:
            continue
        
        time = ws.cell(row, COL_TIME).value
        area = ws.cell(row, COL_AREA).value
        qa_atual = ws.cell(row, COL_QA_PRINCIPAL).value
        dev_atual = ws.cell(row, COL_DEV_PRINCIPAL).value
        
        # Se já está preenchido, pula
        if qa_atual and dev_atual:
            ja_preenchidas += 1
            continue
        
        # Se não tem Time ou Área, não consegue mapear
        if not time or not area:
            sem_mapeamento.append((key, time or "(vazio)", area or "(vazio)"))
            continue
        
        # Busca mapeamento
        chave = (time, area)
        if chave in mapeamento:
            qa_principal = mapeamento[chave]['qa_principal']
            dev_principal = mapeamento[chave]['dev_principal']
            
            # Preenche apenas se estiver vazio
            if not qa_atual:
                ws.cell(row, COL_QA_PRINCIPAL).value = qa_principal
            if not dev_atual:
                ws.cell(row, COL_DEV_PRINCIPAL).value = dev_principal
            
            atualizadas += 1
            print(f"   ✅ {key}: {time} > {area} → QA: {qa_principal}, Dev: {dev_principal}")
        else:
            sem_mapeamento.append((key, time, area))
            print(f"   ⚠️  {key}: {time} > {area} - não encontrado no config")
    
    # 4. Salvar planilha
    print(f"\n💾 Salvando planilha...")
    wb.save(EXCEL_FILE)
    print(f"   ✅ Planilha atualizada")
    wb.close()
    
    # 5. Relatório
    print("\n" + "="*70)
    print("📊 RESUMO DO PREENCHIMENTO")
    print("="*70)
    total = ws.max_row - 1
    print(f"  Total de issues: {total}")
    print(f"  ✅ Atualizadas: {atualizadas}")
    print(f"  ℹ️  Já preenchidas: {ja_preenchidas}")
    print(f"  ⚠️  Sem mapeamento: {len(sem_mapeamento)}")
    
    if sem_mapeamento:
        print(f"\n⚠️  Issues sem mapeamento (Time ou Área não encontrada):")
        for key, time, area in sem_mapeamento[:10]:  # Mostra primeiras 10
            print(f"     • {key}: {time} > {area}")
        if len(sem_mapeamento) > 10:
            print(f"     ... e mais {len(sem_mapeamento) - 10}")
    
    print("="*70)
    print("\n✅ Processo concluído!")
    print(f"   Abra {EXCEL_FILE} para visualizar as alterações.")


# =============================================================================
# EXECUÇÃO
# =============================================================================

if __name__ == "__main__":
    preencher_qa_dev()
