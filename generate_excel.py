"""
RCA Pocket - Gerador de Excel
==============================
Gera (ou atualiza) o arquivo RCA_Pocket.xlsx com 4 abas:
  📊 Dados                   — Issues ativas do Jira + colunas de análise manual
  🗂️ Acompanhamento Issue    — Itens para resolução definitiva
  📦 Arquivo                  — Issues arquivadas (Analisado=Sim + mais de X dias)
  👥 Responsáveis             — Tabela de referência dos times

Se o arquivo já existir:
  - Rows com 'Analisado = Sim' são preservadas integralmente (não sobrescritas pelo Jira)
  - Colunas manuais (K, M–Q) são preservadas nas demais rows
  - Aba Acompanhamento é preservada; novos itens de 'Item p/ Resolução Def.' são adicionados
  - Issues com Analisado=Sim e importadas há mais de dias_retencao vão para aba 📦 Arquivo
"""

import os
import json
from pathlib import Path
from datetime import datetime, timedelta

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

from jira_client import JiraClient, MOCK_ACTIONS
from indexar_testes import carregar_indice, buscar_tas_relacionados
from config_loader import load_config as load_project_config

# =============================================================================
# PALETA DE CORES
# =============================================================================
HEADER_BG       = "1F4E79"   # azul escuro
HEADER_FG       = "FFFFFF"
SUBHEADER_BG    = "2E75B6"   # azul médio
ALT_ROW_BG      = "EBF3FB"   # azul claro
WHITE           = "FFFFFF"

PRIO_CRITICA    = "FF0000"   # vermelho
PRIO_ALTA       = "FF6600"   # laranja
PRIO_MEDIA      = "FFCC00"   # amarelo
PRIO_BAIXA      = "92D050"   # verde

STATUS_RESOLVIDO    = "E2EFDA"  # verde claro
STATUS_EM_ANALISE   = "FFF2CC"  # amarelo claro
STATUS_ABERTO       = "FCE4D6"  # laranja claro

TIPO_PREVENTIVA  = "D9EAD3"   # verde
TIPO_REMEDIACAO  = "FDE9D9"   # laranja


# =============================================================================
# HELPERS DE ESTILO
# =============================================================================

def _make_thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _opaque_color(color: str) -> str:
    """Normaliza cores RGB para ARGB opaco.

    O Excel Online fica inconsistente quando o fill é salvo com alpha 00.
    """
    color = (color or "").strip().replace("#", "").upper()
    if len(color) == 6:
        return f"FF{color}"
    return color


def _solid_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=_opaque_color(color))


def _header_style(ws, row, cols: list, bg=HEADER_BG, fg=HEADER_FG, bold=True):
    for col_idx, val in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = val
        cell.font = Font(bold=bold, color=fg, size=10)
        cell.fill = _solid_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _make_thin_border()


def _set_col_widths(ws, widths: dict):
    """widths: {col_letter_or_index: width}"""
    for col, width in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = width


def _apply_alt_rows(ws, start_row: int, end_row: int, n_cols: int):
    for row in range(start_row, end_row + 1):
        bg = ALT_ROW_BG if row % 2 == 0 else WHITE
        fill = _solid_fill(bg)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE, ALT_ROW_BG):
                cell.fill = fill
            cell.border = _make_thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _add_table(ws, table_name: str, ref: str, style="TableStyleMedium2"):
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)


def _to_excel_date(val):
    """Converte string ISO, string DD/MM/YYYY ou datetime/date para objeto datetime.
    Necessário para que fórmulas FILTER() possam comparar datas corretamente."""
    if val is None or val == "":
        return None
    if hasattr(val, "strftime"):
        # Já é datetime ou date — garante que é datetime
        return val if isinstance(val, datetime) else datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        val = val.strip()
        for fmt, length in [
            ("%Y-%m-%dT%H:%M:%S.%f", 26),
            ("%Y-%m-%dT%H:%M:%S",    19),
            ("%Y-%m-%d",             10),
            ("%d/%m/%Y",             10),
        ]:
            try:
                return datetime.strptime(val[:length], fmt)
            except (ValueError, IndexError):
                continue
    return None


# =============================================================================
# ABA: DADOS
# =============================================================================

DADOS_COLS = [
    # ── BLOCO 1: Identificação Jira (cabeçalho azul escuro A–J) ──────────────
    ("A", "Key",                       14),
    ("B", "Resumo",                    52),
    ("C", "Status Jira",               15),
    ("D", "Prioridade",                13),
    ("E", "Data Criação",              14),
    ("F", "Data Resolução",            14),
    ("G", "Dias p/ Resolver",          14),
    ("H", "DeV Responsável pelo Bug",  20),
    ("I", "QA Responsável pelo Bug",   20),
    ("J", "Qtd Vínculos",              13),
    ("K", "Causa Raiz",                40),
    # ── BLOCO 2: Categorização Jira (cabeçalho slate L–O) ────────────────────
    ("L", "Time",                      16),
    ("M", "Área",                      18),
    ("N", "Tipo Erro (Auto)",          18),
    ("O", "Ação Realizada no Bug",     50),
    # ── BLOCO 3: Análise Manual (cabeçalho verde P–AB) ───────────────────────
    ("P", "Análise da Causa",          40),
    ("Q", "Tipo de Ajuste",            22),
    ("R", "Possui TA",                 14),
    ("S", "Arquivo TA",                35),  # NOVO: referência aos arquivos .robot
    ("T", "Resultado da Automação",    20),
    ("U", "Contexto",                  40),
    ("V", "Problema Resolvido?",       18),
    ("W", "QA Principal",              18),
    ("X", "Dev Principal",             18),
    ("Y", "Issue de Acompanhamento",   18),
    ("Z", "Plano Ação/Lição Aprendida", 32),
    ("AA", "Analisado",                14),
    ("AB", "Data Filtragem",           14),
]

# Cores dos 3 blocos de cabeçalho
_BG_BLOCO1 = "1F4E79"   # azul escuro  — A–K (dados Jira + vínculos + causa raiz)
_BG_BLOCO2 = "44546A"   # slate        — L–O (categorização)
_BG_BLOCO3 = "375623"   # verde escuro — P–AB (análise manual)

def _build_dados(ws, issues: list, tipos_erro: list, preserved: dict):
    ws.title = "📊 Dados"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    # ── Cabeçalho com 3 blocos de cores ──────────────────────────────────────
    headers = [col[1] for col in DADOS_COLS]
    _header_style(ws, 1, headers, bg=_BG_BLOCO1)
    for ci in range(12, 16):   # L–O → bloco 2
        cell = ws.cell(row=1, column=ci)
        cell.fill = _solid_fill(_BG_BLOCO2)
    for ci in range(16, 29):  # P–AB → bloco 3
        cell = ws.cell(row=1, column=ci)
        cell.fill = _solid_fill(_BG_BLOCO3)

    _set_col_widths(ws, {col[0]: col[2] for col in DADOS_COLS})

    # ── Fills pastel por bloco (aplicados em TODAS as linhas de dados) ────────
    fill_b1     = _solid_fill("D6E4F0")  # azul pastel    — A–K
    fill_b2     = _solid_fill("E8EAED")  # cinza-azul pastel — L–O
    fill_b3     = _solid_fill("EBF1DE")  # verde pastel   — P–AB
    fill_frozen = _solid_fill("EFF6FF")  # azul claro — linhas congeladas

    # ── Validações — string inline com unicode explícito (compatível Excel + LibreOffice) ──────
    _sim_nao = '"Sim,N\u00e3o"'   # "Sim,Não"
    _ajuste  = '"C\u00f3digo,Banco de Dados,Infraestrutura,Terceiros"'
    _resultado_auto = '"Detectou problema,N\u00e3o detectou,N\u00e3o se Aplica"'
    dv_ajuste         = DataValidation(type="list", formula1=_ajuste,          allow_blank=True)
    dv_possui_ta      = DataValidation(type="list", formula1=_sim_nao,         allow_blank=True)
    dv_resultado_auto = DataValidation(type="list", formula1=_resultado_auto,  allow_blank=True)
    dv_prob_res       = DataValidation(type="list", formula1=_sim_nao,         allow_blank=True)
    dv_analisado      = DataValidation(type="list", formula1=_sim_nao,         allow_blank=True)

    dados_analisados  = preserved.get("dados_analisados", {})
    dados_manual_cols = preserved.get("dados_manual_cols", {})

    # Carrega índice de TAs para matching por similaridade
    ta_indice = carregar_indice()

    for row_idx, issue in enumerate(issues, start=2):
        key  = issue.get("key", "")
        link = issue.get("link_jira", "")
        ws.row_dimensions[row_idx].height = 20

        # ── Row congelada (Analisado = Sim) ───────────────────────────────────
        if key in dados_analisados:
            preserved_row = dados_analisados[key]
            for ci, col_def in enumerate(DADOS_COLS, start=1):
                val  = preserved_row.get(col_def[1])
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.border    = _make_thin_border()
                cell.font      = Font(size=9)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.fill      = fill_frozen
            kc = ws.cell(row=row_idx, column=1)
            if link:
                kc.hyperlink = link
            kc.font = Font(color="0563C1", underline="single", size=9, bold=True)
            continue

        # ── Row normal ────────────────────────────────────────────────────────
        manual = dados_manual_cols.get(key, {})
        # Campos manuais: preferência para dados preservados do Excel;
        # fallback para campos mock do issue (demo sem Excel prévio)
        def _m(field):
            v = manual.get(field)
            if v:
                return v
            return issue.get(field, "") or ""

        data_criacao   = issue.get("data_criacao")
        data_resolucao = issue.get("data_resolucao")

        vals = [
            key,                                            # A  Key
            issue.get("resumo", ""),                       # B  Resumo
            issue.get("status", ""),                       # C  Status Jira
            issue.get("prioridade", ""),                   # D  Prioridade
            _to_excel_date(data_criacao),                  # E  Data Criação
            _to_excel_date(data_resolucao),                # F  Data Resolução
            None,                                          # G  Dias p/ Resolver ← fórmula
            issue.get("dev_responsavel_bug", ""),          # H  DeV Responsável pelo Bug
            issue.get("qa_responsavel_bug", ""),           # I  QA Responsável pelo Bug
            issue.get("qtd_vinculos", 0),                  # J  Qtd Vínculos
            _m("causa_raiz"),                              # K  Causa Raiz ← manual
            issue.get("time", ""),                         # L  Time
            issue.get("area", ""),                         # M  Área
            issue.get("tipo_erro_auto", ""),               # N  Tipo Erro (Auto)
            issue.get("acao_realizada", ""),               # O  Ação Realizada no Bug
            _m("analise_causa"),                           # P  Análise da Causa ← manual
            _m("ajuste_realizado"),                        # Q  Tipo de Ajuste ← manual
            _m("possui_ta"),                               # R  Possui TA ← manual
            _m("arquivo_ta"),                              # S  Arquivo TA ← automático (validação)
            _m("resultado_automacao"),                     # T  Resultado da Automação ← manual
            _m("contexto"),                                # U  Contexto ← manual
            _m("problema_resolvido"),                      # V  Problema Resolvido? ← manual
            issue.get("qa_principal", ""),                 # W  QA Principal
            issue.get("dev_principal", ""),                # X  Dev Principal
            _m("issue_acompanhamento"),                    # Y  Issue de Acompanhamento ← manual
            _m("plano_acao_licao_aprendida"),              # Z  Plano Ação/Lição Aprendida ← manual
            _m("analisado"),                               # AA Analisado ← manual
            _to_excel_date(issue.get("_data_filtragem", "")),  # AB Data da filtragem
        ]

        # Escreve valores base
        for ci, val in enumerate(vals, start=1):
            cell           = ws.cell(row=row_idx, column=ci, value=val)
            cell.border    = _make_thin_border()
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Aplica cor pastel de bloco em TODAS as células da linha
        for ci in range(1, 12):   ws.cell(row=row_idx, column=ci).fill = fill_b1  # A-K
        for ci in range(12, 16):  ws.cell(row=row_idx, column=ci).fill = fill_b2  # L-O
        for ci in range(16, 29):  ws.cell(row=row_idx, column=ci).fill = fill_b3  # P-AB

        # G (7): fórmula — diferença em dias entre datas
        g_cell = ws.cell(row=row_idx, column=7)
        g_cell.value         = f'=IF(OR(E{row_idx}="",F{row_idx}=""),"",INT(F{row_idx}-E{row_idx}))'
        g_cell.number_format = "0"
        g_cell.fill          = fill_b1

        # A (1): Key como hyperlink
        kc = ws.cell(row=row_idx, column=1)
        if link:
            kc.hyperlink = link
        kc.font = Font(color="0563C1", underline="single", size=9, bold=True)

        # E (5), F (6) e AB (28): formato de data
        for dc in [5, 6, 28]:
            c = ws.cell(row=row_idx, column=dc)
            if c.value:
                c.number_format = "DD/MM/YYYY"

        # O (15): Ação Realizada — wrap text; altura variável
        o_cell = ws.cell(row=row_idx, column=15)
        o_cell.alignment = Alignment(vertical="top", wrap_text=True)
        if vals[14]:  # vals[14] = Ação Realizada (posição O)
            ws.row_dimensions[row_idx].height = 40

        # D (4): cor por prioridade (override do bloco)
        prio_cell   = ws.cell(row=row_idx, column=4)
        colors_prio = {"Crítica": PRIO_CRITICA, "Alta": PRIO_ALTA,
                       "Média": PRIO_MEDIA, "Baixa": PRIO_BAIXA}
        prio_color = colors_prio.get(issue.get("prioridade", ""))
        if prio_color:
            is_dark = issue.get("prioridade") in ("Crítica", "Alta")
            prio_cell.fill = _solid_fill(prio_color)
            prio_cell.font = Font(bold=True, size=9,
                                  color="FFFFFF" if is_dark else "333333")

        # C (3): cor por status (override do bloco)
        status_cell   = ws.cell(row=row_idx, column=3)
        colors_status = {"Resolvido": STATUS_RESOLVIDO,
                         "Em Análise": STATUS_EM_ANALISE, "Aberto": STATUS_ABERTO}
        st_color = colors_status.get(issue.get("status", ""))
        if st_color:
            status_cell.fill = _solid_fill(st_color)

        # AA (27): Data da filtragem/sync
        semana_cell = ws.cell(row=row_idx, column=27)
        semana_cell.alignment = Alignment(vertical="center", horizontal="center")
        semana_cell.font = Font(size=9)
        
        # J (10): Qtd Vínculos - destaque para issues com muitos vínculos
        vinc_cell = ws.cell(row=row_idx, column=10)
        qtd_vinc = issue.get("qtd_vinculos", 0)
        if qtd_vinc >= 5:
            vinc_cell.fill = _solid_fill("FFC7CE")  # vermelho claro
            vinc_cell.font = Font(bold=True, size=9, color="9C0006")
        elif qtd_vinc >= 3:
            vinc_cell.fill = _solid_fill("FFEB9C")  # amarelo
            vinc_cell.font = Font(bold=True, size=9, color="9C6500")

        # S (19): Arquivo TA — preenchido por matching de similaridade
        if ta_indice and not manual.get("arquivo_ta"):
            resumo = issue.get("resumo", "")
            area = issue.get("area", "")
            matches = buscar_tas_relacionados(
                resumo,
                area,
                ta_indice,
                acao_realizada=issue.get("acao_realizada", ""),
                causa_raiz=_m("causa_raiz"),
                analise_causa=_m("analise_causa"),
                contexto=_m("contexto"),
                top_n=5,
            )
            if matches:
                ta_cell = ws.cell(row=row_idx, column=19)
                nomes = [m["nome"] for m in matches[:3]]
                ta_text = "\n".join(nomes)
                if len(matches) > 3:
                    ta_text += f"\n(+{len(matches)-3} mais)"
                ta_cell.value = ta_text
                ta_cell.alignment = Alignment(vertical="top", wrap_text=True)
                # Preenche R (18): Possui TA = "Sim" se encontrou matches
                r_cell = ws.cell(row=row_idx, column=18)
                if not r_cell.value or str(r_cell.value).strip().lower() != "sim":
                    r_cell.value = "Sim"

    n_rows   = len(issues)
    last_row = max(1 + n_rows, 2)
    last_col = len(DADOS_COLS)
    sfx      = last_row + 50

    dv_ajuste.sqref         = "Q2:Q{sfx}".format(sfx=sfx)
    dv_possui_ta.sqref      = "R2:R{sfx}".format(sfx=sfx)
    dv_resultado_auto.sqref = "T2:T{sfx}".format(sfx=sfx)
    dv_prob_res.sqref       = "V2:V{sfx}".format(sfx=sfx)
    dv_analisado.sqref      = "AA2:AA{sfx}".format(sfx=sfx)
    ws.add_data_validation(dv_ajuste)
    ws.add_data_validation(dv_possui_ta)
    ws.add_data_validation(dv_resultado_auto)
    ws.add_data_validation(dv_prob_res)
    ws.add_data_validation(dv_analisado)

    _add_table(ws, "TabelaDados", f"A1:{get_column_letter(last_col)}{last_row}", "TableStyleMedium2")



# =============================================================================
# ABA: ACOMPANHAMENTO ISSUE
# =============================================================================

ACOMP_COLS = [
    ("A", "Issue Acompanhamento", 18),
    ("B", "Issue Original",       16),
    ("C", "Responsável",          22),
    ("D", "Área",                  22),
    ("E", "Ação",                 45),
    ("F", "Status da Ação",       18),
    ("G", "Data Limite",          14),
    ("H", "Data Conclusão",       14),
    ("I", "Observação",           40),
]


def _build_acompanhamento(ws, issues: list, preserved_acomp: list, preserved_manual: dict):
    ws.title = "🗂️ Acompanhamento Issue"
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 32

    headers = [col[1] for col in ACOMP_COLS]
    _header_style(ws, 1, headers, bg=SUBHEADER_BG)
    _set_col_widths(ws, {col[0]: col[2] for col in ACOMP_COLS})

    dv_status = DataValidation(
        type="list",
        formula1='"An\u00e1lise,Andamento,Bloqueado,Conclu\u00eddo"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_status)

    dv_area = DataValidation(
        type="list",
        formula1='"FFC,FatInt,SupCrmImp,RC"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_area)

    # Lookup preservado por Issue de Acompanhamento (dados manuais entre gerações)
    preserved_by_acomp: dict = {r.get("issue_acomp", ""): r for r in preserved_acomp if r.get("issue_acomp")}

    colors_acomp = {
        "Concluído": STATUS_RESOLVIDO,
        "Andamento": STATUS_EM_ANALISE,
        "Análise":   ALT_ROW_BG,
        "Bloqueado": STATUS_ABERTO,
    }

    # Filtra apenas issues que têm Issue de Acompanhamento preenchida (col Y)
    row_idx = 2
    for idx, issue in enumerate(issues):
        dados_row = idx + 2
        key       = issue.get("key", "")
        link      = issue.get("link_jira", "")
        
        # Pega Issue de Acompanhamento da coluna Y (preservada ou mock)
        manual = preserved_manual.get(key, {})
        issue_acomp = manual.get("issue_acompanhamento", "") or issue.get("issue_acompanhamento", "")
        
        # Pula se não tem Issue de Acompanhamento
        if not issue_acomp or not str(issue_acomp).strip():
            continue
        
        issue_acomp = str(issue_acomp).strip()
        acomp_data = preserved_by_acomp.get(issue_acomp, {})
        # Fallback para campos mock
        if not acomp_data:
            acomp_data = {
                "responsavel":    issue.get("acomp_responsavel", ""),
                "area":           issue.get("acomp_area", ""),
                "acao":           issue.get("acomp_acao", ""),
                "status_acao":    issue.get("acomp_status_acao", ""),
                "data_limite":    issue.get("acomp_data_limite"),
                "data_conclusao": issue.get("acomp_data_conclusao"),
                "observacao":     "",
            }

        ws.row_dimensions[row_idx].height = 20

        # A: Issue de Acompanhamento (Key principal desta aba)
        cell_acomp = ws.cell(row=row_idx, column=1, value=issue_acomp)
        cell_acomp.border    = _make_thin_border()
        cell_acomp.font      = Font(size=9, color="0563C1", underline="single", bold=True)
        cell_acomp.alignment = Alignment(vertical="center")
        # TODO: adicionar hyperlink para issue acompanhamento se houver

        # B: Issue Original (hyperlink)
        cell_orig = ws.cell(row=row_idx, column=2, value=key)
        cell_orig.border    = _make_thin_border()
        cell_orig.font      = Font(size=9, color="0563C1", underline="single")
        cell_orig.alignment = Alignment(vertical="center")
        if link:
            cell_orig.hyperlink = link

        # C–I: dados manuais (preservados entre gerações)
        status_val = acomp_data.get("status_acao", "")
        vals_manual = [
            acomp_data.get("responsavel", ""),           # C
            acomp_data.get("area", ""),                  # D
            acomp_data.get("acao", ""),                  # E
            status_val,                                   # F
            (_to_excel_date(acomp_data["data_limite"])
             if acomp_data.get("data_limite") else None),     # G
            (_to_excel_date(acomp_data["data_conclusao"])
             if acomp_data.get("data_conclusao") else None),  # H
            acomp_data.get("observacao", ""),                 # I
        ]
        for ci, val in enumerate(vals_manual, start=3):
            cell           = ws.cell(row=row_idx, column=ci, value=val)
            cell.border    = _make_thin_border()
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 9))

        # Datas (G/H) — formato de data
        for dc in [7, 8]:
            dcf = ws.cell(row=row_idx, column=dc)
            if dcf.value:
                dcf.number_format = "DD/MM/YYYY"

        # Cor por Status (F=6)
        sc = colors_acomp.get(status_val)
        if sc:
            ws.cell(row=row_idx, column=6).fill = _solid_fill(sc)
        
        row_idx += 1

    n_rows   = row_idx - 2
    last_row = max(row_idx - 1, 2)
    last_col = len(ACOMP_COLS)

    dv_status.sqref = f"F2:F{last_row + 50}"
    dv_area.sqref   = f"D2:D{last_row + 50}"

    if n_rows > 0:
        _add_table(ws, "TabelaAcompanhamento",
                   f"A1:{get_column_letter(last_col)}{last_row}", "TableStyleMedium7")

    for extra in range(last_row + 1, last_row + 12):
        ws.row_dimensions[extra].height = 20
        for col in range(1, last_col + 1):
            ws.cell(row=extra, column=col).border = _make_thin_border()
    
    return n_rows  # Retorna número de linhas geradas


# =============================================================================
# ABA: RESPONSÁVEIS
# =============================================================================

def _build_responsaveis(ws, times_config: dict):
    ws.title = "👥 Responsáveis"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    headers = ["Time", "Área", "QA Principal", "QA Secundário", "Dev Principal", "Dev Secundário"]
    _header_style(ws, 1, headers, bg=SUBHEADER_BG)
    _set_col_widths(ws, {"A": 18, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22})

    row_idx = 2
    for time_name, time_data in times_config.items():
        for area in time_data.get("areas", []):
            vals = [
                time_name,
                area.get("nome", ""),
                area.get("qa_principal", ""),
                area.get("qa_secundario", ""),
                area.get("dev_principal", ""),
                area.get("dev_secundario", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _make_thin_border()
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical="center")

            bg = ALT_ROW_BG if row_idx % 2 == 0 else WHITE
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).fill = _solid_fill(bg)

            row_idx += 1

    if row_idx > 2:
        _add_table(ws, "TabelaResponsaveis", f"A1:F{row_idx - 1}", "TableStyleLight9")

    # Instruções
    ws.cell(row=row_idx + 1, column=1).value = "ℹ️ Edite o arquivo rca_config.yaml para alterar responsáveis. Re-execute generate_excel.py para atualizar."
    ws.cell(row=row_idx + 1, column=1).font = Font(italic=True, color="888888", size=8)
    ws.merge_cells(f"A{row_idx + 1}:F{row_idx + 1}")


# =============================================================================
# PRESERVAÇÃO DE DADOS MANUAIS
# =============================================================================

def _read_existing_manual_data(filepath: str) -> dict:
    """
    Lê dados existentes do Excel para preservar entre atualizações.

    Retorna:
        {
          "dados_analisados":  {key: {col_name: val, ...}},  # linhas com Analisado=Sim (row inteira)
          "dados_manual_cols": {key: {col_name: val, ...}},  # colunas manuais de todas as linhas
          "acompanhamento":    [{item, key, data_criacao, responsavel, acao,
                                 status_acao, data_conclusao, observacao}, ...],
        }
    Colunas manuais da aba Dados são localizadas por nome de cabeçalho
    para preservar compatibilidade mesmo quando a posição muda.
    """
    result: dict = {
        "dados_analisados":  {},
        "dados_manual_cols": {},
        "acompanhamento":    [],
    }
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        issues_list: list = []  # ordem das issues para lookup posicional do acompanhamento

        # ── Aba Dados ────────────────────────────────────────────────────────
        dados_sheet = "📊 Dados"
        if dados_sheet in wb.sheetnames:
            ws = wb[dados_sheet]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                wb.close()
                return result

            header = [str(h).strip() if h is not None else "" for h in all_rows[0]]
            # índices 0-based
            col_key       = header.index("Key")            if "Key"                    in header else 0
            col_analisado = header.index("Analisado")      if "Analisado"              in header else None
            col_causa_raiz = header.index("Causa Raiz")         if "Causa Raiz"              in header else None
            col_analise   = header.index("Análise da Causa")    if "Análise da Causa"    in header else None
            col_ajuste    = header.index("Tipo de Ajuste")        if "Tipo de Ajuste"        in header else None
            col_ta        = header.index("Possui TA")              if "Possui TA"              in header else None
            col_resultado_auto = header.index("Resultado da Automação") if "Resultado da Automação" in header else None
            col_contexto  = header.index("Contexto")               if "Contexto"                in header else None
            col_prob      = header.index("Problema Resolvido?")    if "Problema Resolvido?"    in header else None
            col_issue_acomp = header.index("Issue de Acompanhamento") if "Issue de Acompanhamento" in header else None
            col_plano_acao = header.index("Plano Ação/Lição Aprendida") if "Plano Ação/Lição Aprendida" in header else None
            col_semana    = None
            for col_name_semana in ("Data Filtragem", "Semana"):
                if col_name_semana in header:
                    col_semana = header.index(col_name_semana)
                    break
            col_arquivo_ta = header.index("Arquivo TA")               if "Arquivo TA"               in header else None

            manual_indices = {
                "causa_raiz":         col_causa_raiz,
                "analise_causa":      col_analise,
                "ajuste_realizado":   col_ajuste,
                "possui_ta":          col_ta,
                "arquivo_ta":         col_arquivo_ta,
                "resultado_automacao": col_resultado_auto,
                "contexto":           col_contexto,
                "problema_resolvido": col_prob,
                "issue_acompanhamento": col_issue_acomp,
                "plano_acao_licao_aprendida": col_plano_acao,
                "analisado":          col_analisado,
                "semana":             col_semana,
            }

            for row in all_rows[1:]:
                if not any(v for v in row):
                    continue
                key = str(row[col_key] or "").strip()
                if not key:
                    continue

                analisado_val = ""
                if col_analisado is not None and col_analisado < len(row):
                    analisado_val = str(row[col_analisado] or "").strip()

                # Preserva colunas manuais para todas as linhas
                manual_data: dict = {}
                for field_name, ci in manual_indices.items():
                    if ci is not None and ci < len(row):
                        manual_data[field_name] = row[ci]

                result["dados_manual_cols"][key] = manual_data
                issues_list.append({"key": key})

                # Linha inteira preservada se Analisado = "Sim"
                if analisado_val.lower() == "sim":
                    full_row: dict = {}
                    for col_name, ci_val in zip(header, row):
                        full_row[col_name] = ci_val
                    result["dados_analisados"][key] = full_row

        # ── Aba Acompanhamento Issue ──────────────────────────────────────────
        # Col A = Issue Acompanhamento (key desta aba)
        # Col B = Issue Original, Col C = fórmula
        # Preservamos D–I (dados manuais), identificados pela Issue Acomp em A
        acomp_sheet = "🗂️ Acompanhamento Issue"
        if acomp_sheet in wb.sheetnames:
            ws = wb[acomp_sheet]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                # A (índice 0) contém a Issue de Acompanhamento
                issue_acomp = row[0] if len(row) > 0 else None
                if not issue_acomp:
                    continue
                issue_acomp_str = str(issue_acomp).strip()
                if not issue_acomp_str or issue_acomp_str.startswith("="):
                    continue
                result["acompanhamento"].append({
                    "issue_acomp":    issue_acomp_str,
                    "responsavel":    row[3] or "" if len(row) > 3 else "",
                    "area":           row[4] or "" if len(row) > 4 else "",
                    "acao":           row[5] or "" if len(row) > 5 else "",
                    "status_acao":    row[6] or "" if len(row) > 6 else "",
                    "data_limite":    row[7] or None if len(row) > 7 else None,
                    "data_conclusao": row[8] or None if len(row) > 8 else None,
                    "observacao":     row[9] or "" if len(row) > 9 else "",
                })

        wb.close()
    except Exception as e:
        print(f"[WARN] Não foi possível ler dados existentes: {e}")

    return result


# =============================================================================
# DATA FILTRAGEM & ARQUIVAMENTO
# =============================================================================

def _injetar_data_filtragem(issues: list, preserved: dict):
    """
    Injeta em cada issue o campo '_data_filtragem' baseado no valor preservado
    do Excel anterior (coluna "Data Filtragem" / "Semana").
    
    - Issues que já existiam na planilha: mantêm a data original (ex: 14/03/2026)
    - Issues novas (sem dados preservados): recebem a data de hoje
    
    Isso garante que a ordenação por data reflita quando cada issue APARECEU
    pela primeira vez na planilha, não quando o sync rodou.
    """
    hoje_str = datetime.now().strftime("%d/%m/%Y")
    dados_manual = preserved.get("dados_manual_cols", {})
    
    for issue in issues:
        key = issue.get("key", "")
        manual = dados_manual.get(key, {})
        
        # Busca data preservada do Excel (campo "semana" na leitura)
        data_preservada = manual.get("semana")
        
        # Ignora valores legados ("Atual", "Anterior")
        if data_preservada and str(data_preservada).strip().lower() in ("atual", "anterior"):
            data_preservada = None
        
        # Se tem data preservada válida, usa; senão é issue nova → hoje
        if data_preservada and str(data_preservada).strip():
            issue["_data_filtragem"] = str(data_preservada).strip()
        else:
            issue["_data_filtragem"] = hoje_str


def _parse_data_filtragem(data_str) -> datetime:
    """Converte string DD/MM/YYYY para datetime. Fallback: data mínima."""
    if not data_str:
        return datetime(2000, 1, 1)
    if isinstance(data_str, datetime):
        return data_str
    try:
        return datetime.strptime(str(data_str).strip(), "%d/%m/%Y")
    except (ValueError, TypeError):
        return datetime(2000, 1, 1)


def _separar_arquivadas(issues: list, preserved: dict, dias_retencao: int) -> tuple:
    """
    Separa issues em (ativas, arquivadas).
    
    Critério para arquivar:
      - Analisado = "Sim" (no Excel preservado)
      - _data_filtragem > dias_retencao dias atrás
    
    Issues sem Analisado=Sim NUNCA são arquivadas.
    """
    hoje = datetime.now()
    limite = hoje - timedelta(days=dias_retencao)
    dados_analisados = preserved.get("dados_analisados", {})
    
    ativas = []
    arquivadas = []
    
    for issue in issues:
        key = issue.get("key", "")
        
        # Só arquiva se Analisado = Sim
        if key not in dados_analisados:
            ativas.append(issue)
            continue
        
        # Verifica idade pela _data_filtragem (data real de quando apareceu na planilha)
        data_filt = _parse_data_filtragem(issue.get("_data_filtragem"))
        
        if data_filt < limite:
            arquivadas.append(issue)
        else:
            ativas.append(issue)
    
    return ativas, arquivadas


# =============================================================================
# ABA: ARQUIVO
# =============================================================================

ARQUIVO_COLS = [
    ("A", "Key",                       14),
    ("B", "Resumo",                    52),
    ("C", "Status Jira",               15),
    ("D", "Prioridade",                13),
    ("E", "Data Criação",              14),
    ("F", "Data Resolução",            14),
    ("G", "Qtd Vínculos",              13),
    ("H", "Causa Raiz",                40),
    ("I", "Time",                      16),
    ("J", "Tipo Erro (Auto)",          18),
    ("K", "Análise da Causa",          40),
    ("L", "Tipo de Ajuste",            22),
    ("M", "Problema Resolvido?",       18),
    ("N", "Data Filtragem",            14),
    ("O", "Data Arquivamento",         16),
]

_BG_ARQUIVO = "7F6000"  # marrom escuro


def _build_arquivo(ws, issues_arquivadas: list, preserved: dict):
    """Constrói aba 📦 Arquivo com issues que já foram analisadas e são antigas."""
    ws.title = "📦 Arquivo"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    headers = [col[1] for col in ARQUIVO_COLS]
    _header_style(ws, 1, headers, bg=_BG_ARQUIVO)
    _set_col_widths(ws, {col[0]: col[2] for col in ARQUIVO_COLS})

    fill_arq = _solid_fill("FFF8E1")  # amarelo claro

    dados_analisados = preserved.get("dados_analisados", {})
    dados_manual     = preserved.get("dados_manual_cols", {})
    hoje_str = datetime.now().strftime("%d/%m/%Y")

    for row_idx, issue in enumerate(issues_arquivadas, start=2):
        key = issue.get("key", "")
        link = issue.get("link_jira", "")
        ws.row_dimensions[row_idx].height = 20

        # Dados preservados do Excel (linha congelada) ou do issue
        row_data = dados_analisados.get(key, {})
        manual   = dados_manual.get(key, {})

        def _v(excel_col_name, issue_field=None, manual_field=None):
            """Busca valor: primeiro no row preservado, depois manual, depois issue."""
            if row_data and excel_col_name in row_data:
                return row_data[excel_col_name]
            if manual_field and manual.get(manual_field):
                return manual[manual_field]
            if issue_field:
                return issue.get(issue_field, "")
            return ""

        vals = [
            key,                                                  # A Key
            _v("Resumo", "resumo"),                              # B Resumo
            _v("Status Jira", "status"),                         # C Status
            _v("Prioridade", "prioridade"),                      # D Prioridade
            _to_excel_date(_v("Data Criação", "data_criacao")),  # E Data Criação
            _to_excel_date(_v("Data Resolução", "data_resolucao")),  # F Data Resolução
            _v("Qtd Vínculos", "qtd_vinculos"),                  # G Qtd Vínculos
            _v("Causa Raiz", None, "causa_raiz"),                # H Causa Raiz
            _v("Time", "time"),                                  # I Time
            _v("Tipo Erro (Auto)", "tipo_erro_auto"),            # J Tipo Erro
            _v("Análise da Causa", None, "analise_causa"),       # K Análise
            _v("Tipo de Ajuste", None, "ajuste_realizado"),      # L Tipo Ajuste
            _v("Problema Resolvido?", None, "problema_resolvido"),  # M Problema Resolvido
            _to_excel_date(_v("Data Filtragem") or issue.get("_data_filtragem", "")),  # N Data Filtragem
            _to_excel_date(hoje_str),                             # O Data Arquivamento
        ]

        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.border    = _make_thin_border()
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.fill      = fill_arq

        # A: Key como hyperlink
        kc = ws.cell(row=row_idx, column=1)
        if link:
            kc.hyperlink = link
        kc.font = Font(color="0563C1", underline="single", size=9, bold=True)

        # E, F, N, O: formato de data
        for dc in [5, 6, 14, 15]:
            c = ws.cell(row=row_idx, column=dc)
            if c.value:
                c.number_format = "DD/MM/YYYY"

    n_rows = len(issues_arquivadas)
    last_row = max(1 + n_rows, 2)

    if n_rows > 0:
        _add_table(ws, "TabelaArquivo",
                   f"A1:{get_column_letter(len(ARQUIVO_COLS))}{last_row}",
                   "TableStyleMedium4")

    # Nota informativa
    info_row = last_row + 2
    ws.cell(row=info_row, column=1).value = (
        f"ℹ️ Issues com Analisado=Sim importadas há mais de "
        f"{preserved.get('_dias_retencao', 30)} dias são movidas para esta aba automaticamente."
    )
    ws.cell(row=info_row, column=1).font = Font(italic=True, color="888888", size=8)
    ws.merge_cells(f"A{info_row}:{get_column_letter(len(ARQUIVO_COLS))}{info_row}")


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def _sort_issues_by_priority(issues: list) -> list:
    """
    Ordena issues com dupla ordenação:
    
    🔴 1º CRITÉRIO: Data Filtragem (mais recente primeiro)
       Cada nova importação "empilha" sobre as anteriores no topo da planilha.
       Usa _data_filtragem (preservada do Excel anterior) para manter consistência.
    
    🟠 2º CRITÉRIO: Quantidade de Vínculos (mais vínculos primeiro)
       Dentro do mesmo lote de importação, issues com mais vínculos sobem.
    
    🟡 3º CRITÉRIO: Criticidade (P0 > P1 > outras)
       Desempate dentro do mesmo lote e mesma qtd de vínculos.
    
    Exemplo com 2 lotes:
      Lote 16/03 (mais recente — topo):
        1. Issue com 10 vínculos + P1
        2. Issue com 5 vínculos + P0
        3. Issue com 0 vínculos + P1
      Lote 14/03 (anterior — abaixo):
        4. Issue com 8 vínculos + P0
        5. Issue com 2 vínculos + P1
    
    Retorna nova lista ordenada.
    """
    def _sort_key(issue):
        qtd_vinculos = issue.get("qtd_vinculos", 0)
        prioridade = issue.get("prioridade", "")
        
        # Usa _data_filtragem (data preservada do Excel, não data_importacao do sync)
        data_filt = _parse_data_filtragem(issue.get("_data_filtragem"))
        dia_filtragem = data_filt.date()
        
        # Peso da prioridade: Crítica=0, Alta=1, demais=2
        if prioridade == "Crítica":
            peso_prio = 0
        elif prioridade == "Alta":
            peso_prio = 1
        else:
            peso_prio = 2
        
        # 1º Data Filtragem DESC (dia mais recente primeiro)
        # 2º Vínculos DESC (mais vínculos primeiro)
        # 3º Prioridade ASC (P0 > P1 > outras)
        return (-dia_filtragem.toordinal(), -qtd_vinculos, peso_prio)
    
    return sorted(issues, key=_sort_key)


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def generate_excel(config: dict, output_path: str = None):
    """
    Gera o arquivo RCA_Pocket.xlsx com 4 abas:
      📊 Dados               — issues ativas do Jira + colunas manuais
      🗂️ Acompanhamento Issue — itens de resolução definitiva + plano de ação
      📦 Arquivo              — issues arquivadas (Analisado=Sim + >dias_retencao dias)
      👥 Responsáveis         — mapeamento time → QA/Dev

    Preservação entre re-gerações:
      - Linhas com Analisado = "Sim" ficam intocadas (frozen).
      - Colunas manuais (K, M–Q) são preservadas para demais linhas.
      - Aba Acompanhamento Issue é totalmente preservada; novos itens de
        "Item p/ Resolução Def." são acrescentados sem duplicar.
    """
    output_path = output_path or config.get("excel", {}).get("arquivo_saida", "RCA_Pocket.xlsx")
    output_path = Path(output_path)

    print(f"\n{'='*60}")
    print(f"  RCA Pocket — Gerador de Excel")
    print(f"{'='*60}")

    # 1. Buscar issues no Jira
    client = JiraClient(config)
    issues = client.get_normalized_issues()
    print(f"✅ {len(issues)} issues carregadas")

    # 2. Preservar dados manuais se arquivo já existe
    preserved: dict = {"dados_analisados": {}, "dados_manual_cols": {}, "acompanhamento": []}
    if output_path.exists():
        print(f"🔄 Arquivo existente detectado — preservando dados manuais...")
        preserved = _read_existing_manual_data(str(output_path))
        n_frozen = len(preserved["dados_analisados"])
        n_acomp  = len(preserved["acompanhamento"])
        print(f"   ↳ {n_frozen} linhas congeladas (Analisado=Sim) | "
              f"{n_acomp} itens de acompanhamento preservados")
    
    # 2.5. Injetar "Data Filtragem" preservada em cada issue (para ordenação e exibição)
    _injetar_data_filtragem(issues, preserved)

    # 2.6. Arquivamento: separar issues analisadas com mais de X dias
    dias_retencao = config.get("excel", {}).get("dias_retencao", 30)
    issues_ativas, issues_arquivadas = _separar_arquivadas(
        issues, preserved, dias_retencao
    )
    preserved["_dias_retencao"] = dias_retencao
    if issues_arquivadas:
        print(f"📦 {len(issues_arquivadas)} issues arquivadas (Analisado=Sim + >{dias_retencao} dias)")
    
    # 2.7. Ordenação (só nas ativas)
    print(f"🔀 Ordenando issues (data filtragem > vínculos > prioridade)...")
    issues_ativas = _sort_issues_by_priority(issues_ativas)
    
    # Contagem por data
    hoje_str = datetime.now().strftime("%d/%m/%Y")
    n_hoje = sum(1 for i in issues_ativas if i.get("_data_filtragem") == hoje_str)
    print(f"   ↳ {n_hoje} issues novas (hoje) | {len(issues_ativas) - n_hoje} anteriores")

    # Tipos de erro disponíveis (para dropdown)
    tipos_erro = list(config.get("tipos_erro", {}).keys())

    # 3. Construir workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]


    ws_dados = wb.create_sheet("📊 Dados")
    ws_acomp = wb.create_sheet("🗂️ Acompanhamento Issue")
    ws_arq   = wb.create_sheet("📦 Arquivo")
    ws_resp  = wb.create_sheet("👥 Responsáveis")

    print("📊 Construindo aba Dados...")
    _build_dados(ws_dados, issues_ativas, tipos_erro, preserved)

    print("📦 Construindo aba Arquivo...")
    _build_arquivo(ws_arq, issues_arquivadas, preserved)

    print("🗂️  Construindo aba Acompanhamento Issue...")
    n_acomp_gerados = _build_acompanhamento(ws_acomp, issues_ativas, preserved["acompanhamento"], preserved["dados_manual_cols"])

    print("👥 Construindo aba Responsáveis...")
    _build_responsaveis(ws_resp, config.get("times", {}))

    # 4. Metadados
    wb.properties.title = "RCA Pocket"
    wb.properties.description = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # 5. Salvar
    wb.save(str(output_path))
    print(f"\n✅ Excel gerado: {output_path.resolve()}")
    print(f"   Abas: {' | '.join(wb.sheetnames)}")
    print(f"   Issues ativas: {len(issues_ativas)} | Arquivadas: {len(issues_arquivadas)} | Acompanhamento: {n_acomp_gerados} itens")

    # Salvar também no OneDrive, se configurado
    onedrive_path = config.get("excel", {}).get("onedrive_path", "")
    if onedrive_path:
        try:
            wb.save(str(onedrive_path))
            print(f"\n☁️  Cópia salva no OneDrive: {onedrive_path}")
        except Exception as e:
            print(f"[WARN] Falha ao salvar no OneDrive: {e}")

    print(f"{'='*60}\n")
    return str(output_path.resolve())


if __name__ == "__main__":
    cfg = load_project_config(__file__)
    generate_excel(cfg)
