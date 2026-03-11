#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Preenche Time/Área usando inferência inteligente baseada em:
1. Análise do resumo e descrição da issue
2. Padrões conhecidos de palavras-chave
3. Histórico de classificações
"""

import os
import json
import yaml
import openpyxl
import re

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================
CACHE_FILE = "data/issues_cache.json"
EXCEL_FILE = "RCA_Pocket.xlsx"
CONFIG_FILE = "rca_config.yaml"
SHEET_NAME = "📊 Dados"

COL_KEY = 1     # A
COL_RESUMO = 2  # B
COL_TIME = 12   # L
COL_AREA = 13   # M

# =============================================================================
# MAPEAMENTO POR PALAVRAS-CHAVE
# =============================================================================

KEYWORDS_TIME_AREA = {
    "Suprimentos": {
        "Entrada XML": ["xml", "nfe", "nota fiscal", "entrada", "importação xml", "importacao"],
        "Balanço": ["balanço", "balanco", "inventário", "inventario", "contagem"],
        "Compras 2.0": ["compras", "pedido de compra", "sugestão de compra", "sugestao"],
        "Estoque": ["estoque", "movimentação", "movimentacao", "saldo"]
    },
    "FFC": {
        "Conciliadores": ["conciliador", "conciliação", "conciliacao"],
        "Gestão Financeira": ["gestão financeira", "gestao financeira", "financeiro"],
        "NF-e": ["nfe", "nf-e", "nota fiscal eletrônica"],
        "Rejeições": ["rejeição", "rejeicao", "rejeições", "rejeicoes", "sefaz"]
    },
    "FatInt": {
        "Venda Fácil": ["venda fácil", "venda facil", "pdv", "pos"],
        "B2C": ["b2c", "e-commerce", "ecommerce", "loja virtual"],
        "B2B": ["b2b"],
        "Integração": ["integração", "integracao", "webhook", "api"]
    }
}

# =============================================================================
# FUNÇÕES
# =============================================================================

def carregar_cache():
    """Carrega cache de issues"""
    with open(CACHE_FILE, 'r', encoding='utf-8') as f:
        cache = json.load(f)
    return cache.get('issues', [])

def inferir_time_area_por_conteudo(resumo, descricao=""):
    """
    Infere Time e Área analisando resumo e descrição.
    
    Returns:
        Tupla (time, area, confiança) ou (None, None, 0)
    """
    texto = (resumo + " " + descricao).lower()
    
    melhor_match = None
    melhor_score = 0
    
    for time, areas in KEYWORDS_TIME_AREA.items():
        for area, keywords in areas.items():
            score = 0
            for keyword in keywords:
                if keyword in texto:
                    score += 1
            
            if score > melhor_score:
                melhor_score = score
                melhor_match = (time, area)
    
    if melhor_match and melhor_score > 0:
        return (*melhor_match, melhor_score)
    
    # Fallback: se menciona "Suprimentos" no texto mas nenhuma área específica
    if "suprimentos" in texto or "suprimento" in texto:
        return ("Suprimentos", "Estoque", 0.5)
    
    if "faturamento" in texto or "ffc" in texto:
        return ("FFC", "Gestão Financeira", 0.5)
    
    if "venda" in texto or "pdv" in texto:
        return ("FatInt", "Venda Fácil", 0.5)
    
    return (None, None, 0)

def preencher_com_inferencia():
    """Preenche Time/Área usando inferência inteligente"""
    print("\n" + "="*70)
    print("  🧠 PREENCHIMENTO INTELIGENTE DE TIME E ÁREA")
    print("="*70)
    
    # Carregar dados
    print("\n📦 Carregando cache...")
    issues = carregar_cache()
    issues_by_key = {i['key']: i for i in issues}
    
    print(f"   ✅ {len(issues)} issues no cache")
    
    # Abrir planilha
    print(f"\n📄 Abrindo planilha: {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    atualizadas = 0
    baixa_confianca = []
    sem_classificacao = []
    
    print(f"\n🔍 Processando issues com inferência inteligente...")
    
    for row_idx in range(2, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=COL_KEY)
        key = key_cell.value
        
        if not key:
            continue
        
        # Buscar issue no cache
        issue = issues_by_key.get(key)
        if not issue:
            continue
        
        # Tentar inferir por conteúdo
        resumo = issue.get('resumo', '')
        descricao = issue.get('descricao', '')
        
        time, area, confianca = inferir_time_area_por_conteudo(resumo, descricao)
        
        # Atualizar planilha
        time_cell = ws.cell(row=row_idx, column=COL_TIME)
        area_cell = ws.cell(row=row_idx, column=COL_AREA)
        
        if time and confianca > 0:
            time_cell.value = time
            area_cell.value = area
            atualizadas += 1
            
            if confianca >= 2:
                print(f"   ✅ {key}: {time} > {area} (confiança: alta ✓✓)")
            elif confianca >= 1:
                print(f"   ⚠️  {key}: {time} > {area} (confiança: média ✓)")
                baixa_confianca.append(key)
            else:
                print(f"   ⚠️  {key}: {time} > {area} (confiança: baixa ~)")
                baixa_confianca.append(key)
        else:
            sem_classificacao.append({
                'key': key,
                'resumo': resumo[:60]
            })
            print(f"   ❌ {key}: sem classificação - \"{resumo[:50]}...\"")
    
    # Salvar planilha
    print(f"\n💾 Salvando planilha...")
    wb.save(EXCEL_FILE)
    print(f"   ✅ Planilha atualizada")
    
    # Resumo
    print("\n" + "="*70)
    print("📊 RESUMO DO PREENCHIMENTO INTELIGENTE")
    print("="*70)
    print(f"  Total: {ws.max_row - 1} issues")
    print(f"  ✅ Classificadas: {atualizadas}")
    print(f"  ⚠️  Confiança baixa/média: {len(baixa_confianca)} (revisar manualmente)")
    print(f"  ❌ Não classificadas: {len(sem_classificacao)}")
    
    if baixa_confianca:
        print(f"\n⚠️  Revise manualmente as issues com confiança baixa/média:")
        for k in baixa_confianca[:5]:
            print(f"     • {k}")
        if len(baixa_confianca) > 5:
            print(f"     ... e mais {len(baixa_confianca) - 5}")
    
    if sem_classificacao:
        print(f"\n❌ Issues não classificadas (adicione keywords ou classifique manualmente):")
        for item in sem_classificacao[:5]:
            print(f"     • {item['key']}: \"{item['resumo']}\"")
        if len(sem_classificacao) > 5:
            print(f"     ... e mais {len(sem_classificacao) - 5}")
    
    print("="*70)
    return atualizadas > 0

# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("\n╔══════════════════════════════════════════════════════╗")
    print("║  RCA Pocket - Preenchimento Inteligente Time/Área   ║")
    print("╚══════════════════════════════════════════════════════╝")
    
    try:
        sucesso = preencher_com_inferencia()
        
        if sucesso:
            print("\n✅ Processo concluído!")
            print("   Abra RCA_Pocket.xlsx para revisar as classificações.\n")
        else:
            print("\n⚠️  Nenhuma issue foi classificada automaticamente")
            print("   Considere adicionar mais palavras-chave no script")
            print("   ou preencher manualmente as colunas Time e Área.\n")
    
    except KeyboardInterrupt:
        print("\n\n⚠️  Processo cancelado")
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
