#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa dados da planilha de exemplo (RCA_Pocket_Exemplo.xlsx) para a planilha oficial.
Aplica toda a lógica de preenchimento automático, ordenação e formatação.
"""

import openpyxl
import json
import os
from datetime import datetime
from pathlib import Path


def ler_planilha_exemplo():
    """Lê a planilha de exemplo e extrai dados das abas Dados e Acompanhamento."""
    print("\n📥 Lendo planilha de exemplo (RCA_Pocket_Exemplo.xlsx)...")
    
    try:
        wb = openpyxl.load_workbook("RCA_Pocket_Exemplo.xlsx", data_only=True)
    except FileNotFoundError:
        print("❌ Arquivo RCA_Pocket_Exemplo.xlsx não encontrado!")
        return None, None
    
    # Lê aba Dados
    ws_dados = wb["📊 Dados"]
    issues = []
    
    # Mapeamento de nomes de colunas para chaves internas
    mapeamento = {
        "key": "key",
        "resumo": "summary",
        "status_jira": "status",
        "prioridade": "priority",
        "data_criação": "created",
        "data_resolução": "resolutiondate",
        "dias_p/_resolver": "tempo_resolucao_dias",
        "dev_responsável_pelo_bug": "dev_responsavel_bug",
        "qa_responsável_pelo_bug": "qa_responsavel_bug",
        "qtd_vínculos": "qtd_vinculos",
        "causa_raiz": "causa_raiz",
        "time": "time",
        "área": "area",
        "tipo_de_erro": "tipo_erro_auto",
        "revisar_classificação": "revisar_classificacao",
        "tipo_de_erro_(manual)": "tipo_erro_manual",
        "ação_realizada_no_bug": "acao_realizada",
        "análise_da_causa": "analise_causa",
        "ajuste_realizado": "ajuste_realizado",
        "possui_ta": "possui_ta",
        "problema_resolvido": "problema_resolvido",
        "qa_principal": "qa_principal",
        "dev_principal": "dev_principal",
        "issue_de_acompanhamento": "issue_acomp",
        "analisado": "analisado",
        "semana": "semana",
    }
    
    # Cabeçalho está na linha 1, dados começam na linha 2
    for row_idx in range(2, ws_dados.max_row + 1):
        issue_dict = {}
        tem_dados = False
        
        # Lê cada coluna
        for col_idx in range(1, 27):  # 26 colunas (A-Z)
            cell = ws_dados.cell(row_idx, col_idx)
            valor = cell.value
            
            # Pega nome da coluna do cabeçalho
            col_nome = ws_dados.cell(1, col_idx).value
            if col_nome:
                col_key = col_nome.lower().replace(" ", "_")
                issue_dict[col_key] = valor
                
                if valor is not None and valor != "":
                    tem_dados = True
        
        if tem_dados and issue_dict.get("key"):
            issues.append(issue_dict)
    
    print(f"   ✅ {len(issues)} issues lidas da aba Dados")
    
    # Lê aba Acompanhamento
    acompanhamento = []
    if "🗂️ Acompanhamento Issue" in wb.sheetnames:
        ws_acomp = wb["🗂️ Acompanhamento Issue"]
        
        for row_idx in range(2, ws_acomp.max_row + 1):
            item = {}
            tem_dados = False
            
            for col_idx in range(1, 9):  # 8 colunas
                cell = ws_acomp.cell(row_idx, col_idx)
                valor = cell.value
                
                col_nome = ws_acomp.cell(1, col_idx).value
                if col_nome:
                    col_key = col_nome.lower().replace(" ", "_")
                    item[col_key] = valor
                    
                    if valor is not None and valor != "":
                        tem_dados = True
            
            if tem_dados:
                acompanhamento.append(item)
        
        print(f"   ✅ {len(acompanhamento)} itens lidos da aba Acompanhamento")
    
    wb.close()
    return issues, acompanhamento


def preencher_campos_vazios(issues):
    """Preenche campos que estão vazios na planilha exemplo."""
    print("\n🔧 Preenchendo campos vazios...")
    
    preenchidos = 0
    
    for issue in issues:
        alterado = False
        
        # Status Jira: Se vazio, preenche com "Resolvido"
        if not issue.get("status_jira"):
            issue["status_jira"] = "Resolvido"
            alterado = True
        
        # Prioridade: Se vazio, preenche com "Média"
        if not issue.get("prioridade"):
            issue["prioridade"] = "Média"
            alterado = True
        
        # Qtd Vínculos: Se vazio, preenche com 0
        if not issue.get("qtd_vínculos"):
            issue["qtd_vínculos"] = 0
            alterado = True
        
        # Time e Área: Tenta inferir do resumo ou deixa vazio
        # (pode ser implementado no futuro com lógica mais sofisticada)
        if not issue.get("time"):
            issue["time"] = ""
        
        if not issue.get("área"):
            issue["área"] = ""
        
        if alterado:
            preenchidos += 1
    
    print(f"   ✅ {preenchidos} issues com campos preenchidos")
    return issues


def converter_para_formato_jira(issues):
    """Converte issues da planilha exemplo para o formato do jira_client."""
    print("\n🔄 Convertendo dados para formato do Jira...")
    
    issues_jira = []
    
    for issue in issues:
        # Converte para formato esperado pelo generate_excel
        issue_jira = {
            "key": issue.get("key", ""),
            "link_jira": f"https://jira.example.com/browse/{issue.get('key', '')}",
            "resumo": issue.get("resumo", ""),
            "status": issue.get("status_jira", "Resolvido"),
            "prioridade": issue.get("prioridade", "Média"),
            "data_criacao": issue.get("data_criação"),
            "data_resolucao": issue.get("data_resolução"),
            "responsavel_jira": issue.get("responsável_pela_issue_-_jira", ""),
            "dev_responsavel_bug": issue.get("dev_responsável_pelo_bug", ""),
            "qa_responsavel_bug": issue.get("qa_responsável_pelo_bug", ""),
            "qtd_vinculos": issue.get("qtd_vínculos", 0),
            "causa_raiz": issue.get("causa_raiz", ""),
            "time": issue.get("time", ""),
            "area": issue.get("área", ""),
            "tipo_erro_auto": issue.get("tipo_de_erro", ""),
            "revisar_classificacao": issue.get("revisar_classificação", ""),
            "tipo_erro_manual": issue.get("tipo_de_erro_(manual)", ""),
            "acao_realizada": issue.get("ação_realizada_no_bug", ""),
            "analise_causa": issue.get("análise_da_causa", ""),
            "ajuste_realizado": issue.get("ajuste_realizado", ""),
            "possui_ta": issue.get("possui_ta", ""),
            "problema_resolvido": issue.get("problema_resolvido", ""),
            "qa_principal": issue.get("qa_principal", ""),
            "dev_principal": issue.get("dev_principal", ""),
            "issue_acomp": issue.get("issue_de_acompanhamento", ""),
            "analisado": issue.get("analisado", ""),
            "semana": issue.get("semana", ""),
        }
        
        issues_jira.append(issue_jira)
    
    print(f"   ✅ {len(issues_jira)} issues convertidas")
    return issues_jira


def ordenar_issues(issues):
    """Aplica ordenação: vínculos > P0 > P1 > data."""
    print("\n🔀 Aplicando ordenação (vínculos > P0 > P1)...")
    
    # Usa a função de ordenação do generate_excel
    from generate_excel import _sort_issues_by_priority
    issues_ordenadas = _sort_issues_by_priority(issues)
    
    print(f"   ✅ {len(issues_ordenadas)} issues ordenadas")
    return issues_ordenadas


def gerar_planilha_oficial(issues, acompanhamento):
    """Gera a planilha oficial usando generate_excel (com toda a formatação)."""
    print("\n📊 Gerando planilha oficial (RCA_Pocket.xlsx)...")
    
    # Converte issues para o formato normalizado final (não precisa passar por normalize_issue novamente)
    # pois os dados já vêm da planilha no formato esperado
    normalized_issues = []
    
    for issue in issues:
        # Issue já está no formato interno completo, só precisa garantir os campos necessários
        normalized = {
            "key": issue.get("key", ""),
            "link_jira": f"https://jira.example.com/browse/{issue.get('key', '')}",
            "resumo": issue.get("resumo", ""),
            "descricao": "",
            "status": issue.get("status", "Resolvido"),
            "prioridade": issue.get("prioridade", "Média"),
            "data_criacao": issue.get("data_criacao"),
            "data_atualizacao": None,
            "data_resolucao": issue.get("data_resolucao"),
            "tempo_resolucao_dias": None,
            "responsavel_jira": issue.get("responsavel_jira", ""),
            "labels": "",
            "tipo_erro_auto": issue.get("tipo_erro_auto", ""),
            "tipo_erro_manual": issue.get("tipo_erro_manual", ""),
            "acao_realizada": issue.get("acao_realizada", ""),
            "causa_raiz": issue.get("causa_raiz", ""),
            "analisado": issue.get("analisado", ""),
            "revisar_classificacao": issue.get("revisar_classificacao", ""),
            "time": issue.get("time", ""),
            "area": issue.get("area", ""),
            "qa_principal": issue.get("qa_principal", ""),
            "qa_secundario": "",
            "dev_principal": issue.get("dev_principal", ""),
            "dev_secundario": "",
            "dev_responsavel_bug": issue.get("dev_responsavel_bug", ""),
            "qa_responsavel_bug": issue.get("qa_responsavel_bug", ""),
            "qtd_vinculos": issue.get("qtd_vinculos", 0),
            "data_importacao": datetime.now(),
            "analise_causa": issue.get("analise_causa", ""),
            "ajuste_realizado": issue.get("ajuste_realizado", ""),
            "possui_ta": issue.get("possui_ta", ""),
            "problema_resolvido": issue.get("problema_resolvido", ""),
            "acomp_area": "",
            "acomp_responsavel": "",
            "acomp_acao": "",
            "acomp_status_acao": "",
            "acomp_data_conclusao": None,
        }
        
        # Calcula tempo de resolução se houver datas
        if normalized["data_criacao"] and normalized["data_resolucao"]:
            if isinstance(normalized["data_criacao"], datetime) and isinstance(normalized["data_resolucao"], datetime):
                normalized["tempo_resolucao_dias"] = (normalized["data_resolucao"] - normalized["data_criacao"]).days
        
        normalized_issues.append(normalized)
    
    # Salva cache no formato normalizado (não precisa de normalização posterior)
    cache_data = {
        "meta": {
            "total": len(normalized_issues),
            "last_sync": datetime.now().isoformat(),
            "jql": "Imported from example spreadsheet",
            "force_use_cache": True,  # Flag para forçar uso mesmo em modo mock
        },
        "synced_at": datetime.now().isoformat(),
        "total": len(normalized_issues),
        "issues": normalized_issues
    }
    
    # Salva cache (sobrescreve o existente)
    os.makedirs("data", exist_ok=True)
    cache_file = "data/issues_cache.json"
    
    # Backup do cache atual se existir
    if os.path.exists(cache_file):
        import shutil
        backup_file = "data/issues_cache_backup.json"
        shutil.copy(cache_file, backup_file)
        print(f"   📦 Backup do cache original: {backup_file}")
    
    # Converte datas para ISO antes de salvar
    for issue in cache_data["issues"]:
        for key in ["data_criacao", "data_resolucao", "data_atualizacao", "data_importacao"]:
            if key in issue and isinstance(issue[key], datetime):
                issue[key] = issue[key].isoformat()
    
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"   💾 Cache atualizado: {cache_file}")
    
    # Chama generate_excel que lerá do cache
    from generate_excel import generate_excel
    from config_loader import load_config as load_project_config
    
    # Carrega config
    config = load_project_config()
    
    # Gera o Excel (isso aplicará TODA a formatação correta)
    print(f"   🎨 Aplicando formatação completa...")
    generate_excel(config)
    
    print(f"   ✅ Planilha oficial gerada com {len(issues)} issues")


def main():
    print("=" * 80)
    print("  🔄 IMPORTAÇÃO DA PLANILHA DE EXEMPLO PARA OFICIAL")
    print("=" * 80)
    
    # 1. Lê planilha de exemplo
    issues, acompanhamento = ler_planilha_exemplo()
    
    if not issues:
        print("\n❌ Nenhum dado encontrado na planilha de exemplo!")
        return
    
    # 2. Preenche campos vazios
    issues = preencher_campos_vazios(issues)
    
    # 3. Converte para formato Jira
    issues_jira = converter_para_formato_jira(issues)
    
    # 4. Ordena
    issues_ordenadas = ordenar_issues(issues_jira)
    
    # 5. Gera planilha oficial (com toda formatação!)
    gerar_planilha_oficial(issues_ordenadas, acompanhamento or [])
    
    print("\n" + "=" * 80)
    print("✅ IMPORTAÇÃO CONCLUÍDA COM SUCESSO!")
    print("   → Planilha oficial gerada com TODA a formatação aplicada")
    print("=" * 80)
    print()


if __name__ == "__main__":
    main()
