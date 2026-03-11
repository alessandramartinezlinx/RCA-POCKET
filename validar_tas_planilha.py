"""
Validação de Testes Automatizados (TAs) vs. Planilha RCA Pocket

Este script busca no repositório Robot Framework do GitHub se existem
testes automatizados associados aos issues registrados na planilha RCA.

Funcionalidades:
1. Lê issues da planilha RCA_Pocket.xlsx (coluna Key)
2. Busca no GitHub por testes Robot que mencionam cada issue
3. Atualiza coluna "Possui TA" automaticamente
4. Gera relatório de cobertura

Dependências:
- pip install PyGithub openpyxl

Configuração:
- Gerar token GitHub: https://github.com/settings/tokens
- Permissões necessárias: repo (read-only)
"""

import os
import json
import time
from datetime import datetime
from typing import List, Dict, Tuple
import openpyxl
from github import Github, GithubException, RateLimitExceededException

# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

GITHUB_REPO = "MEDIUM-RETAIL-MICROVIX/ta-robotframework"
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")  # Token de acesso ao GitHub
EXCEL_FILE = "RCA_Pocket.xlsx"
SHEET_NAME = "📊 Dados"

# Colunas da planilha (conforme generate_excel.py)
COL_KEY = 1          # A: Jira Key
COL_POSSUI_TA = 18   # R: Possui TA
COL_ARQUIVO_TA = 19  # S: Arquivo TA (referência aos arquivos .robot)

# Cache para evitar múltiplas requisições
CACHE_FILE = "data/ta_validation_cache.json"

# Controle de rate limit
DELAY_ENTRE_BUSCAS = 2.5  # segundos entre cada busca (evita rate limit)


# =============================================================================
# FUNÇÕES AUXILIARES
# =============================================================================

def carregar_cache() -> Dict:
    """Carrega cache de validações anteriores"""
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def salvar_cache(cache: Dict):
    """Salva cache de validações"""
    os.makedirs(os.path.dirname(CACHE_FILE), exist_ok=True)
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


def verificar_rate_limit(github_client: Github):
    """Verifica e aguarda se estiver próximo do rate limit"""
    try:
        rate_limit = github_client.get_rate_limit()
        search_limit = rate_limit.search
        
        if search_limit.remaining < 5:
            reset_time = search_limit.reset
            wait_seconds = (reset_time - datetime.now()).total_seconds() + 10
            if wait_seconds > 0:
                print(f"\n⚠️  Rate limit baixo! Aguardando {int(wait_seconds)}s até reset...")
                time.sleep(wait_seconds)
    except Exception as e:
        # Se falhar ao verificar rate limit, continua (não bloqueia)
        pass


def buscar_ta_para_issue(github_client: Github, issue_key: str, cache: Dict) -> Tuple[bool, List[str]]:
    """
    Busca no GitHub se existe TA para o issue especificado
    
    Retorna:
        (possui_ta: bool, arquivos_encontrados: List[str])
    """
    # Verifica cache
    if issue_key in cache:
        print(f"  [CACHE] {issue_key}: {cache[issue_key]['possui_ta']}")
        return cache[issue_key]['possui_ta'], cache[issue_key].get('arquivos', [])
    
    print(f"  Buscando TAs para {issue_key}...", end='', flush=True)
    
    try:
        # Busca COMBINADA: (MODAJOI-XXXXX OR SHOP-JOI-XXXXX) em arquivos .robot
        # Reduz de 2 requisições para 1 por issue
        shop_key = issue_key.replace("MODAJOI-", "SHOP-JOI-")
        query = f"repo:{GITHUB_REPO} {issue_key} OR {shop_key} extension:robot"
        
        results = github_client.search_code(query, order='desc')
        
        arquivos_encontrados = []
        
        # Itera com segurança - verifica totalCount primeiro
        if results.totalCount > 0:
            for i, result in enumerate(results):
                if i >= 10:  # Limita a 10 primeiros
                    break
                if result.path.endswith('.robot'):
                    arquivos_encontrados.append(result.path)
        
        possui_ta = len(arquivos_encontrados) > 0
        
        # Salva no cache
        cache[issue_key] = {
            'possui_ta': possui_ta,
            'arquivos': arquivos_encontrados,
            'data_consulta': datetime.now().isoformat()
        }
        
        if possui_ta:
            print(f" ✅ {len(arquivos_encontrados)} TA(s)")
        else:
            print(f" ❌ Nenhum TA")
        
        return possui_ta, arquivos_encontrados
    
    except RateLimitExceededException:
        print(f" ⏳ Rate limit atingido")
        # Não faz retry aqui - deixa o loop principal lidar
        return False, []
    
    except GithubException as e:
        if e.status == 403:
            print(f" ⚠️  Erro 403 - Sem acesso")
        else:
            print(f" ⚠️  Erro {e.status}")
        return False, []
    
    except Exception as e:
        print(f" ⚠️  Erro: {type(e).__name__}")
        return False, []


def ler_issues_da_planilha() -> List[str]:
    """Lê lista de issues (Keys) da planilha RCA Pocket"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    issues = []
    for row in range(2, ws.max_row + 1):  # Pula header
        key = ws.cell(row, COL_KEY).value
        if key and isinstance(key, str) and key.startswith("MODAJOI-"):
            issues.append(key)
    
    wb.close()
    return issues


def atualizar_coluna_possui_ta(resultados: Dict[str, Tuple[bool, List[str]]]):
    """
    Atualiza colunas 'Possui TA' e 'Arquivo TA' na planilha com resultados da validação
    
    Args:
        resultados: Dict {issue_key: (possui_ta, [arquivos])}
    """
    print("\n📝 Atualizando planilha...")
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    atualizados = 0
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row, COL_KEY).value
        if key in resultados:
            possui_ta, arquivos = resultados[key]
            valor_novo = "Sim" if possui_ta else "Não"
            valor_atual = ws.cell(row, COL_POSSUI_TA).value
            
            if valor_atual != valor_novo:
                ws.cell(row, COL_POSSUI_TA).value = valor_novo
                atualizados += 1
            
            # Preencher coluna Arquivo TA com os arquivos encontrados
            if possui_ta and arquivos:
                # Pegar só os nomes dos arquivos (sem path completo)
                nomes_arquivos = [arq.split('/')[-1] for arq in arquivos[:3]]
                arquivo_ta_text = ", ".join(nomes_arquivos)
                if len(arquivos) > 3:
                    arquivo_ta_text += f" (+{len(arquivos)-3})"
                ws.cell(row, COL_ARQUIVO_TA).value = arquivo_ta_text
            else:
                ws.cell(row, COL_ARQUIVO_TA).value = ""
    
    if atualizados > 0:
        wb.save(EXCEL_FILE)
        print(f"  ✅ {atualizados} issue(s) atualizado(s)")
    else:
        print("  ℹ️  Nenhuma atualização necessária")
    
    wb.close()


def gerar_relatorio(resultados: Dict[str, Tuple[bool, List[str]]]):
    """Gera relatório de cobertura de TAs"""
    total = len(resultados)
    com_ta = sum(1 for possui_ta, _ in resultados.values() if possui_ta)
    sem_ta = total - com_ta
    
    print("\n" + "="*70)
    print("📊 RELATÓRIO DE COBERTURA DE TESTES AUTOMATIZADOS")
    print("="*70)
    print(f"\nTotal de issues analisados: {total}")
    print(f"  ✅ Com TA:  {com_ta} ({com_ta/total*100:.1f}%)")
    print(f"  ❌ Sem TA:  {sem_ta} ({sem_ta/total*100:.1f}%)")
    
    # Detalhamento de issues COM TA
    if com_ta > 0:
        print("\n" + "-"*70)
        print("Issues COM teste automatizado:")
        print("-"*70)
        for issue_key, (possui_ta, arquivos) in resultados.items():
            if possui_ta:
                print(f"\n  {issue_key}:")
                for arquivo in arquivos[:3]:  # Máximo 3 arquivos
                    print(f"    • {arquivo}")
                if len(arquivos) > 3:
                    print(f"    ... e mais {len(arquivos)-3} arquivo(s)")
    
    # Sugestões de issues SEM TA (priorizar por prioridade/data)
    if sem_ta > 0:
        print("\n" + "-"*70)
        print("⚠️  Issues SEM teste automatizado (sugestão de priorização):")
        print("-"*70)
        for issue_key, (possui_ta, _) in resultados.items():
            if not possui_ta:
                print(f"  • {issue_key}")
    
    print("\n" + "="*70)


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================

def validar_tas_planilha():
    """Função principal de validação"""
    print("🔍 VALIDAÇÃO DE TESTES AUTOMATIZADOS - RCA POCKET")
    print("="*70)
    
    # 1. Verificar token GitHub
    if not GITHUB_TOKEN:
        print("\n⚠️  ATENÇÃO: Token GitHub não configurado!")
        print("Configure a variável de ambiente GITHUB_TOKEN")
        print("Instruções: https://github.com/settings/tokens")
        return
    
    github_client = Github(GITHUB_TOKEN)
    
    # 2. Carregar cache
    cache = carregar_cache()
    print(f"\n📦 Cache carregado: {len(cache)} issue(s) em cache")
    
    # 3. Ler issues da planilha
    print(f"\n📄 Lendo planilha {EXCEL_FILE}...")
    issues = ler_issues_da_planilha()
    print(f"  {len(issues)} issue(s) encontrado(s)\n")
    
    # 4. Buscar TAs para cada issue
    print("🔎 Iniciando busca de TAs no GitHub...")
    print(f"⏱️  Estimativa: ~{len(issues) * DELAY_ENTRE_BUSCAS:.0f}s ({len(issues)} issues × {DELAY_ENTRE_BUSCAS}s delay)")
    
    # Verificação inicial de rate limit
    verificar_rate_limit(github_client)
    
    resultados = {}
    for i, issue_key in enumerate(issues, start=1):
        # Verifica rate limit antes de cada busca
        verificar_rate_limit(github_client)
        
        possui_ta, arquivos = buscar_ta_para_issue(github_client, issue_key, cache)
        resultados[issue_key] = (possui_ta, arquivos)
        
        # Delay APÓS cada busca para respeitar rate limit
        if i < len(issues):  # Não precisa delay após última busca
            time.sleep(DELAY_ENTRE_BUSCAS)
        
        # A cada 10 issues: salva checkpoint
        if i % 10 == 0:
            print(f"\n💾 Checkpoint: {i}/{len(issues)} issues processados")
            salvar_cache(cache)
    
    # 5. Salvar cache atualizado (final)
    salvar_cache(cache)
    print(f"\n✅ Busca concluída: {len(resultados)} issues processados")
    
    # 6. Atualizar planilha (passa resultados completos com arquivos)
    atualizar_coluna_possui_ta(resultados)
    
    # 7. Gerar relatório
    gerar_relatorio(resultados)
    
    print("\n✅ Validação concluída!")


# =============================================================================
# EXECUÇÃO
# =============================================================================

if __name__ == "__main__":
    validar_tas_planilha()
